import datetime
import os
import pandas as pd
from base_utils import rel2fullpath
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import collections
import seaborn as sns
from definitions import SOURCE_DIR


def load_member_data(relative_excel_path: str, year_to_sheet: collections.defaultdict):
    input_file = rel2fullpath(relative_excel_path)

    df = pd.read_excel(input_file, sheet_name=year_to_sheet[year])
    new_columns = [column.lower().replace('.', '').replace(' ', '_').replace('/', '_').replace(':', '')
                  for column in df.columns]
    df.columns = [column.lower().replace('å', 'a').replace('ä', 'a').replace('ö', 'o')
                  for column in new_columns]

    df = df.assign(birth_date=[datetime.datetime.strptime(day, '%Y-%m-%d') for day in df.fodelsedat_personnr])
    today = datetime.datetime(year, 12, 31)
    print(today)
    df = df.assign(age=[relativedelta(today, birth).years for birth in df.birth_date])

    # Binned age
    age_splits = [-1, 7, 12, 16, 20, 25, 100000]
    labels = ['0-7', '8-12', '13-16', '17-20', '21-25', '26+']
    bin_category_to_number = {}
    for key, item in enumerate(labels):
        bin_category_to_number[item] = key
    print(bin_category_to_number)

    df = df.assign(binned=pd.cut(df['age'], bins=age_splits, labels=labels))
    df = df.assign(binned=[str(age_bin) for age_bin in df.binned])
    df = df.assign(bin_order=df.binned.replace(bin_category_to_number))
    return df, bin_category_to_number


def add_frame_to_axis(ax):
    for position in ["left", "bottom", "right", "top"]:
        ax.spines[position].set_visible(True)
        ax.spines[position].set_color('0.5')


def add_labels(ax):
    for c in ax.containers:
        labels = [(v.get_x() + v.get_width()/2, int(v.get_height())) for v in c]
        for (x, y) in labels:
            ax.text(x, y, str(y), ha="center", va="bottom")


if __name__ == '__main__':
    pd.set_option('display.max_rows', 500)
    pd.set_option('display.max_columns', 500)
    pd.set_option('display.width', 1000)

    output_directory = f'{SOURCE_DIR}/member_analysis'
    if not os.path.exists(output_directory):
        raise ValueError('Unknown folder ' + output_directory)

    years = [2019, 2020, 2021, 2022, 2023, 2024]
    year_to_sheet_name = collections.defaultdict(lambda: 'SearchPersons')
    year_to_sheet_name[2019] = 'Data'
    excel_flag = False

    summary = pd.DataFrame()
    for year in years:
        excel_file = 'data/members/ExportedPersons_' + str(year) + '.xlsx'
        df, cat_to_num = load_member_data(excel_file, year_to_sheet_name)

        count_df = df[['kon', 'binned', 'fornamn']].groupby(by=['kon', 'binned']).count()
        count_df = count_df.reset_index()
        count_df.columns = ['Kön', 'Ålder', 'Antal']
        count_df = count_df.assign(**{'År': year})
        summary = pd.concat([summary, count_df], axis=0)
        print(summary)

        if excel_flag:
            df_out = df[['fornamn', 'efternamn', 'fodelsedat_personnr', 'kon', 'age', 'binned']]
            print(df_out)

            output_excel = rel2fullpath(os.path.join(os.path.split(excel_file)[0], 'PersonAges_' + str(year) + '.xlsx'))
            wb = openpyxl.Workbook()
            worksheet = wb.create_sheet('Data')
            for r in dataframe_to_rows(df_out, index=False, header=True):
                worksheet.append(r)

            worksheet = wb.create_sheet('Stat')
            for r in dataframe_to_rows(count_df, index=False, header=True):
                worksheet.append(r)

            wb.save(output_excel)

    year_sex_total = summary.groupby(["Kön", 'År']).sum().reset_index(drop=False)

    g = sns.catplot(
        data=year_sex_total, kind="bar",
        x="År", y="Antal", hue="Kön",
        palette="dark", alpha=.6, height=6
    )
    g.despine(left=True)
    g.set_axis_labels("År", "Antal medlemmar")
    g.legend.set_title("")
    g.fig.set_size_inches(7, 4)
    figure = output_directory + '/yearly_overview_since_2019.png'
    ax = g.facet_axis(0, 0)
    add_labels(ax)
    add_frame_to_axis(ax)
    g.savefig(figure, dpi=300)
    print('Created ' + figure)

    latest_year = summary['År'].max()
    this_year = summary.loc[summary['År'] == latest_year]
    this_year = this_year.assign(bin_number=this_year['Ålder'].replace(cat_to_num))
    this_year = this_year.sort_values(by='bin_number')

    this_year = this_year.reset_index(drop=True)
    this_year.to_excel(output_directory + '/this_year.xlsx', index=False)

    g = sns.catplot(
        data=this_year, kind="bar",
        x="Ålder", y="Antal", hue="Kön",
        palette=["Red", "Blue"], alpha=.6, height=6,
    )
    g.despine(left=True)
    g.legend.set_title("")
    g.set_axis_labels("Åldersspann", "Antal medlemmar")
    g.fig.subplots_adjust(top=0.9)  # adjust the Figure in rp
    g.fig.suptitle(f"Totalt antal HSOK-medlemmar: {this_year['Antal'].sum()} ({latest_year}-12-31)")

    ax = g.facet_axis(0, 0)
    add_labels(ax)
    add_frame_to_axis(ax)
    g.fig.set_size_inches(7, 3)

    g.savefig('this_year.png', dpi=300)

    men = this_year.loc[this_year["Kön"] == 'Man'].Antal.sum()
    women = this_year.loc[this_year["Kön"] == 'Kvinna'].Antal.sum()
    total = this_year["Antal"].sum()

    text_result = [str(f"Antalet klubbmedlemmar var {total} vid verksamhetsårets slut. "
                       f"Antalet män {men} st och antalet kvinnor {women} st."), ""]
    for gender in this_year["Kön"].unique():
        gender_data = this_year.loc[this_year["Kön"] == gender]
        if gender == 'Man':
            gender_string = 'Män'
        else:
            gender_string = 'Kvinnor'
        gender_result = []
        for key, row in gender_data.iterrows():
            gender_result.append(f"{gender_string} {row['Ålder']} år: {row['Antal']} st.")
        gender_result = " ".join(gender_result)
        text_result.append(gender_result)

    text_result.append(" ")
    text_result.append(f"Styrelsen tackar alla som gjort en insats för klubben under året och hoppas på ett "
                       f"riktigt bra verksamhetsår {str(this_year['År'].unique()[0] + 1)}.")
    print("\n".join(text_result))
    print("")
#
